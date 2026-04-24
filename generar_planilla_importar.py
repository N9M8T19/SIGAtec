"""
generar_planilla_importar.py
============================
Lee todas las planillas individuales de docentes (formato ET N°7)
y genera una planilla unificada con pestaña IMPORTAR lista para
subir a Google Sheets y usar con SIGA-Tec.

USO:
    python generar_planilla_importar.py

    Por defecto busca los .xlsx en la carpeta ./horarios_docentes/
    y genera el archivo HORARIOS_IMPORTAR_2026.xlsx en la misma carpeta.

    Para cambiar las rutas, editar las constantes al principio del script.
"""

import os
import re
import sys
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURACIÓN — editá estas rutas si es necesario
# ─────────────────────────────────────────────
CARPETA_ENTRADA = Path("./horarios_docentes")   # carpeta con los .xlsx individuales
ARCHIVO_SALIDA  = Path("./HORARIOS_IMPORTAR_2026.xlsx")
NOMBRE_HOJA_ACTIVA = "HORARIO 26"               # nombre de la pestaña a leer en cada planilla

# ─────────────────────────────────────────────
# MAPEO: horario impreso → código de módulo SIGA-Tec
# ─────────────────────────────────────────────
HORARIO_A_MODULO = {
    "07.30 a 08.10": "M1",
    "08.10 a 08.50": "M2",
    "09.00 a 09.40": "M3",
    "09.40 a 10.20": "M4",
    "10.30 a 11.10": "M5",
    "11.10 a 11.50": "M6",
    "12.00 a 12.40": "M7",
    "12.40 a 13.20": "M8",
    "13.20 a 14.00": "T2",
    "14.00 a 14.40": "T3",
    "14.50 a 15.30": "T4",
    "15.30 a 16.10": "T5",
    "16.20 a 17.00": "T6",
    "17.00 a 17.40": "T7",
    "17.40 a 18.20": "T8",
}

# Columnas de inicio de cada día en la grilla (índice 0-based)
# Estructura: Asignatura | Horario | Curso  (3 columnas por día)
DIAS_COLUMNAS = {
    "LUNES":     0,
    "MARTES":    3,
    "MIÉRCOLES": 6,
    "JUEVES":    9,
    "VIERNES":   12,
}

# Celdas que se ignoran (solo las completamente vacías)
IGNORAR_ASIGNATURAS = {""}

# ─────────────────────────────────────────────
# Funciones auxiliares
# ─────────────────────────────────────────────

def normalizar_texto(texto):
    """Limpia espacios y pasa a mayúsculas."""
    if texto is None:
        return ""
    return str(texto).strip().upper()


def normalizar_horario(horario):
    """
    Normaliza el string de horario para que coincida con el mapeo.
    Acepta variantes como '13.20 a 14.00', '13:20 a 14:00', '13.20-14.00'.
    """
    if horario is None:
        return ""
    h = str(horario).strip()
    # Reemplazar ':' por '.' en las horas
    h = re.sub(r'(\d{2}):(\d{2})', r'\1.\2', h)
    # Reemplazar ' - ' o '-' por ' a '
    h = re.sub(r'\s*[-–]\s*', ' a ', h)
    # Eliminar espacios dobles
    h = re.sub(r'\s+', ' ', h)
    return h


def normalizar_curso(curso):
    """
    Convierte 'N 1 G 6', 'N1G6', 'N1 G6', etc. al formato 'N1G6'.
    """
    if curso is None:
        return ""
    c = str(curso).strip().upper()
    # Eliminar espacios entre letra y número
    c = re.sub(r'N\s*(\d+)\s*G\s*(\d+)', r'N\1G\2', c)
    return c


def extraer_nombre_docente(ws):
    """
    Lee el nombre del docente de la planilla.
    El nombre está en la fila 4 (índice 3), pero en la planilla de ejemplo
    la celda A4 dice 'Docente' y el nombre real parece estar en el nombre
    del archivo. Intentamos leerlo de la celda B4 o de la celda que sigue
    al label 'Docente'.
    Si no lo encontramos, devolvemos None para que el llamador use el nombre
    del archivo.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=6, values_only=True))
    for row in rows:
        for i, cell in enumerate(row):
            if normalizar_texto(cell) == "DOCENTE":
                # El nombre puede estar en la misma fila a la derecha
                # o en la siguiente celda no vacía
                for j in range(i + 1, len(row)):
                    val = normalizar_texto(row[j])
                    if val and val not in ("CARGO", "TURNO", "TP1", "TP2", "TP3",
                                           "MAÑANA", "TARDE", "VESPERTINO"):
                        return val
    return None


def nombre_desde_archivo(ruta):
    """
    Extrae el nombre del docente desde el nombre del archivo.
    Ejemplo: 'RUBIO_BRENDA_HORARIO_2026.xlsx' → 'RUBIO BRENDA'
    Toma todo antes de '_HORARIO' o '_horario', o las primeras 2 palabras.
    """
    stem = Path(ruta).stem.upper()
    # Intentar cortar en '_HORARIO'
    if "_HORARIO" in stem:
        nombre = stem[:stem.index("_HORARIO")].replace("_", " ").strip()
        return nombre
    # Fallback: las primeras 2 palabras separadas por _
    partes = stem.split("_")
    return " ".join(partes[:2]).strip()


def parsear_planilla(ruta_archivo):
    """
    Lee una planilla individual y devuelve una lista de dicts:
    [{'apellido_nombre': ..., 'dia': ..., 'modulo': ..., 'materia': ..., 'curso': ...}, ...]
    """
    try:
        wb = load_workbook(ruta_archivo, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ⚠️  No se pudo abrir {ruta_archivo.name}: {e}")
        return []

    # Elegir la hoja correcta
    if NOMBRE_HOJA_ACTIVA in wb.sheetnames:
        ws = wb[NOMBRE_HOJA_ACTIVA]
    else:
        ws = wb.active

    # Leer todas las filas en memoria (evitar problemas con read_only y múltiples iteraciones)
    todas_las_filas = list(ws.iter_rows(values_only=True))

    # Nombre del docente
    nombre_docente = extraer_nombre_docente(ws)
    if not nombre_docente:
        nombre_docente = nombre_desde_archivo(ruta_archivo)

    # Encontrar la fila de los días (contiene 'LUNES')
    fila_dias = None
    fila_headers = None
    for idx, row in enumerate(todas_las_filas):
        vals = [normalizar_texto(c) for c in row]
        if "LUNES" in vals:
            fila_dias = idx
            fila_headers = idx + 1
            break

    if fila_dias is None:
        print(f"  ⚠️  No se encontró la grilla de días en {ruta_archivo.name}")
        wb.close()
        return []

    # Confirmar columnas de inicio de cada día leyendo la fila de días
    dia_cols = {}
    fila_d = todas_las_filas[fila_dias]
    for col_idx, cell_val in enumerate(fila_d):
        dia = normalizar_texto(cell_val)
        if dia in DIAS_COLUMNAS:
            dia_cols[dia] = col_idx

    if not dia_cols:
        print(f"  ⚠️  No se detectaron columnas de días en {ruta_archivo.name}")
        wb.close()
        return []

    # Procesar filas de datos (desde fila_headers + 1 en adelante)
    registros = []
    for row in todas_las_filas[fila_headers + 1:]:
        # Parar si la fila está completamente vacía
        if all(c is None for c in row):
            break

        for dia, col_inicio in dia_cols.items():
            # Cada día tiene 3 columnas: Asignatura, Horario, Curso
            try:
                asignatura = normalizar_texto(row[col_inicio])
                horario    = normalizar_horario(row[col_inicio + 1]) if len(row) > col_inicio + 1 else ""
                curso      = normalizar_curso(row[col_inicio + 2])   if len(row) > col_inicio + 2 else ""
            except IndexError:
                continue

            # Ignorar celdas vacías o extra clases
            if not asignatura or asignatura in IGNORAR_ASIGNATURAS:
                continue

            # Convertir horario a código de módulo
            modulo = HORARIO_A_MODULO.get(horario)
            if not modulo:
                print(f"  ⚠️  Horario no reconocido: '{horario}' "
                      f"(docente: {nombre_docente}, día: {dia}, materia: {asignatura})")
                continue

            registros.append({
                "apellido_nombre": nombre_docente,
                "dia":             dia,
                "modulo":          modulo,
                "materia":         asignatura,
                "curso":           curso,
            })

    wb.close()
    return registros


# ─────────────────────────────────────────────
# Generador de la planilla de salida
# ─────────────────────────────────────────────

def generar_planilla_importar(todos_los_registros, docentes_sin_datos):
    """Crea el archivo Excel con pestaña IMPORTAR."""

    wb = Workbook()
    ws = wb.active
    ws.title = "IMPORTAR"

    # ── Estilos ──
    color_header   = "1A2A6C"   # azul marino (color SIGA-Tec)
    color_acento   = "E8821A"   # naranja (color SIGA-Tec)
    color_fila_par = "EEF1F8"   # azul muy claro para filas pares

    fuente_titulo  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    fuente_header  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    fuente_normal  = Font(name="Calibri", size=10)
    fuente_aviso   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

    relleno_header = PatternFill("solid", fgColor=color_header)
    relleno_acento = PatternFill("solid", fgColor=color_acento)
    relleno_par    = PatternFill("solid", fgColor=color_fila_par)

    centrado       = Alignment(horizontal="center", vertical="center")
    izquierda      = Alignment(horizontal="left",   vertical="center")

    borde_fino = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ── Fila 1: título ──
    ws.merge_cells("A1:E1")
    ws["A1"] = "SIGA-Tec — Importación de Horarios de Docentes 2026"
    ws["A1"].font      = fuente_titulo
    ws["A1"].fill      = relleno_header
    ws["A1"].alignment = centrado
    ws.row_dimensions[1].height = 28

    # ── Fila 2: aviso ──
    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "⚠  No modificar los nombres de columna. "
        "El campo 'apellido_nombre' debe coincidir exactamente con el nombre cargado en SIGA-Tec."
    )
    ws["A2"].font      = fuente_aviso
    ws["A2"].fill      = relleno_acento
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # ── Fila 3: encabezados ──
    COLUMNAS = ["apellido_nombre", "dia", "modulo", "materia", "curso"]
    ANCHOS   = [30, 14, 10, 30, 12]

    for col_idx, (nombre_col, ancho) in enumerate(zip(COLUMNAS, ANCHOS), start=1):
        celda = ws.cell(row=3, column=col_idx, value=nombre_col)
        celda.font      = fuente_header
        celda.fill      = relleno_header
        celda.alignment = centrado
        celda.border    = borde_fino
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[3].height = 18

    # ── Filas de datos ──
    for fila_idx, reg in enumerate(todos_los_registros, start=4):
        relleno_fila = relleno_par if fila_idx % 2 == 0 else None
        for col_idx, campo in enumerate(COLUMNAS, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=reg[campo])
            celda.font      = fuente_normal
            celda.alignment = centrado if col_idx > 1 else izquierda
            celda.border    = borde_fino
            if relleno_fila:
                celda.fill = relleno_fila
        ws.row_dimensions[fila_idx].height = 16

    # ── Fijar encabezados ──
    ws.freeze_panes = "A4"

    # ── Hoja de errores si hay docentes sin datos ──
    if docentes_sin_datos:
        ws_err = wb.create_sheet("ADVERTENCIAS")
        ws_err["A1"] = "Archivo"
        ws_err["B1"] = "Motivo"
        ws_err["A1"].font = fuente_header
        ws_err["B1"].font = fuente_header
        ws_err["A1"].fill = relleno_acento
        ws_err["B1"].fill = relleno_acento
        ws_err.column_dimensions["A"].width = 45
        ws_err.column_dimensions["B"].width = 50
        for i, (archivo, motivo) in enumerate(docentes_sin_datos, start=2):
            ws_err.cell(row=i, column=1, value=archivo)
            ws_err.cell(row=i, column=2, value=motivo)

    wb.save(ARCHIVO_SALIDA)


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  SIGA-Tec — Generador de planilla IMPORTAR")
    print("=" * 60)

    if not CARPETA_ENTRADA.exists():
        print(f"\n❌ No existe la carpeta '{CARPETA_ENTRADA}'.")
        print("   Creala y copiá ahí todas las planillas individuales de docentes.")
        sys.exit(1)

    archivos = sorted(CARPETA_ENTRADA.glob("*.xlsx"))
    if not archivos:
        print(f"\n❌ No se encontraron archivos .xlsx en '{CARPETA_ENTRADA}'.")
        sys.exit(1)

    print(f"\n📂 Carpeta: {CARPETA_ENTRADA.resolve()}")
    print(f"📄 Archivos encontrados: {len(archivos)}\n")

    todos_los_registros = []
    docentes_sin_datos  = []
    docentes_procesados = []

    for archivo in archivos:
        print(f"  → Procesando: {archivo.name}")
        registros = parsear_planilla(archivo)
        if registros:
            nombre = registros[0]["apellido_nombre"]
            print(f"     ✅ {nombre} — {len(registros)} módulos encontrados")
            todos_los_registros.extend(registros)
            docentes_procesados.append(nombre)
        else:
            print(f"     ⚠️  Sin registros válidos")
            docentes_sin_datos.append((archivo.name, "No se encontraron módulos válidos"))

    print()
    print("─" * 60)
    print(f"  Total de docentes procesados : {len(docentes_procesados)}")
    print(f"  Total de módulos encontrados : {len(todos_los_registros)}")
    if docentes_sin_datos:
        print(f"  ⚠️  Archivos con problemas    : {len(docentes_sin_datos)}")
    print("─" * 60)

    if not todos_los_registros:
        print("\n❌ No se generó ningún registro. Revisá los archivos de entrada.")
        sys.exit(1)

    generar_planilla_importar(todos_los_registros, docentes_sin_datos=docentes_sin_datos)

    print(f"\n✅ Planilla generada: {ARCHIVO_SALIDA.resolve()}")
    print()
    print("  PRÓXIMOS PASOS:")
    print("  1. Abrí el archivo y verificá que los nombres en 'apellido_nombre'")
    print("     coincidan exactamente con los cargados en SIGA-Tec.")
    print("  2. Subí el archivo a Google Drive como Google Sheets.")
    print("  3. Renombrá la pestaña a 'IMPORTAR' si no lo está ya.")
    print("  4. Desde SIGA-Tec → Importar → Horarios de Docentes, pegá")
    print("     el ID del spreadsheet y ejecutá la importación.")
    print()


if __name__ == "__main__":
    main()
