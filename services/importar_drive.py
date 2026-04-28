"""
services/importar_drive.py
Importa datos desde Google Sheets usando cuenta de servicio.

Estructura de planilla de carros:
- Fila 1: COD CARRO (encabezado del carro — el número viene del nombre de la pestaña)
- Fila 2: N° INTERNO | N° DE SERIE (encabezados de columnas)
- Fila 3+: datos de netbooks (col A = N° INTERNO, col B = N° DE SERIE)
"""

import re
import os
import json
from datetime import datetime
from models import db, Docente, Netbook, Carro, PantallaDigital, HistorialPantalla, Alumno

SERVICE_ACCOUNT_FILE = 'service_account.json'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']


def _extraer_sheet_id(url):
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if match:
        return match.group(1)
    raise ValueError('URL de Google Sheets inválida.')


def _get_service():
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    # Opción 1: variable de entorno SERVICE_ACCOUNT_JSON (producción en Render)
    sa_json = os.environ.get('SERVICE_ACCOUNT_JSON')
    if sa_json:
        try:
            sa_info = json.loads(sa_json)
            creds = service_account.Credentials.from_service_account_info(
                sa_info, scopes=SCOPES
            )
            return build('sheets', 'v4', credentials=creds)
        except Exception as e:
            raise ValueError(f'Error al leer SERVICE_ACCOUNT_JSON: {e}')

    # Opción 2: archivo local service_account.json (desarrollo local)
    if os.path.exists(SERVICE_ACCOUNT_FILE):
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
        return build('sheets', 'v4', credentials=creds)

    # Ninguna opción disponible
    raise FileNotFoundError(
        'No se encontró la cuenta de servicio. '
        'En producción: configurá la variable de entorno SERVICE_ACCOUNT_JSON en Render. '
        'En local: colocá el archivo service_account.json en la raíz del proyecto.'
    )


def _leer_hoja(service, sheet_id, nombre_hoja):
    result = service.spreadsheets().values().get(
        spreadsheetId=sheet_id,
        range=f"'{nombre_hoja}'!A:Z"
    ).execute()
    return result.get('values', [])


def obtener_pestanas(url):
    sheet_id = _extraer_sheet_id(url)
    service  = _get_service()
    metadata = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
    pestanas = [hoja['properties']['title'] for hoja in metadata.get('sheets', [])]
    return pestanas, sheet_id


def _extraer_numero_carro(nombre_hoja):
    """
    Extrae el número de carro del nombre de la pestaña.
    Ejemplos: 'CARRO 10' → '10', 'CARRO FIJO TALLER AULA 1' → 'FIJO TALLER AULA 1'
    """
    nombre_upper = nombre_hoja.upper().strip()
    if nombre_upper.startswith('CARRO '):
        return nombre_upper.replace('CARRO ', '').strip()
    return nombre_hoja.strip()


# ─────────────────────────────────────────────────────────────────────────────
#  IMPORTAR CARROS Y NETBOOKS
# ─────────────────────────────────────────────────────────────────────────────

def importar_carro_desde_hoja(sheet_id, nombre_hoja):
    """
    Importa un carro y sus netbooks desde una pestaña.

    Estructura esperada:
    - Fila 1: COD CARRO (encabezado — se ignora, el número viene del nombre de pestaña)
    - Fila 2: N° INTERNO | N° DE SERIE (encabezados de columnas — se ignoran)
    - Fila 3+: datos reales (col A = N° INTERNO, col B = N° DE SERIE)
    """
    resultados = {
        'hoja':         nombre_hoja,
        'importados':   0,
        'actualizados': 0,
        'errores':      [],
        'carro':        None
    }

    service = _get_service()
    filas   = _leer_hoja(service, sheet_id, nombre_hoja)

    # Necesitamos al menos 3 filas (2 encabezados + 1 dato)
    if not filas or len(filas) < 3:
        resultados['errores'].append(f'La pestaña "{nombre_hoja}" no tiene datos.')
        return resultados

    # El número de carro viene del nombre de la pestaña
    numero_carro = _extraer_numero_carro(nombre_hoja)
    resultados['carro'] = numero_carro

    # Buscar o crear el carro
    carro = Carro.query.filter_by(numero_fisico=numero_carro).first()
    if not carro:
        carro = Carro(numero_fisico=numero_carro, estado='activo')
        db.session.add(carro)
        db.session.flush()

    # Saltar las 2 primeras filas (encabezados) y leer desde fila 3
    datos = filas[2:]  # índice 2 = fila 3 en la planilla

    for i, fila in enumerate(datos, start=3):
        try:
            if not fila or len(fila) < 1:
                continue

            n_interno = str(fila[0]).strip() if len(fila) > 0 else ''
            n_serie   = str(fila[1]).strip() if len(fila) > 1 else ''

            # Saltar filas vacías
            if not n_interno and not n_serie:
                continue

            # Buscar netbook por número de serie
            netbook = Netbook.query.filter_by(numero_serie=n_serie).first() if n_serie else None

            if netbook:
                # Actualizar existente
                netbook.carro_id       = carro.id
                netbook.numero_interno = n_interno
                resultados['actualizados'] += 1
            else:
                # Crear nueva
                netbook = Netbook(
                    carro_id       = carro.id,
                    numero_interno = n_interno,
                    numero_serie   = n_serie,
                    estado         = 'operativa'
                )
                db.session.add(netbook)
                resultados['importados'] += 1

        except Exception as e:
            resultados['errores'].append(f'Fila {i + 2}: {str(e)}')

    db.session.commit()
    return resultados


# ─────────────────────────────────────────────────────────────────────────────
#  IMPORTAR DOCENTES
# ─────────────────────────────────────────────────────────────────────────────

def importar_docentes(sheet_id, nombre_hoja):
    """
    Columnas: Apellido y Nombre | DNI | Email
    Fila 1 = encabezados, fila 2+ = datos.
    Actualiza por DNI, agrega nuevos.
    """
    resultados = {'hoja': nombre_hoja, 'importados': 0, 'actualizados': 0, 'errores': []}

    service = _get_service()
    filas   = _leer_hoja(service, sheet_id, nombre_hoja)

    if not filas or len(filas) < 2:
        resultados['errores'].append('La pestaña está vacía.')
        return resultados

    for i, fila in enumerate(filas[1:], start=2):
        try:
            if len(fila) < 2:
                continue

            nombre_completo = fila[0].strip() if len(fila) > 0 else ''
            dni             = str(fila[1]).strip() if len(fila) > 1 else ''
            email           = fila[2].strip() if len(fila) > 2 else ''

            if not nombre_completo or not dni:
                continue

            if ',' in nombre_completo:
                partes   = nombre_completo.split(',', 1)
                apellido = partes[0].strip()
                nombre   = partes[1].strip()
            else:
                partes   = nombre_completo.split(' ', 1)
                apellido = partes[0].strip()
                nombre   = partes[1].strip() if len(partes) > 1 else ''

            docente = Docente.query.filter_by(dni=dni).first()
            if docente:
                docente.nombre   = nombre
                docente.apellido = apellido
                docente.correo   = email or docente.correo
                resultados['actualizados'] += 1
            else:
                docente = Docente(
                    dni=dni, nombre=nombre, apellido=apellido,
                    correo=email, activo=True
                )
                db.session.add(docente)
                resultados['importados'] += 1

        except Exception as e:
            resultados['errores'].append(f'Fila {i}: {str(e)}')

    db.session.commit()
    return resultados


# ─────────────────────────────────────────────────────────────────────────────
#  IMPORTAR PANTALLAS DIGITALES
# ─────────────────────────────────────────────────────────────────────────────

def importar_pantallas(sheet_id, nombre_hoja):
    """
    Columnas: AULA | CURSO | NÚMERO DE SERIE
    Fila 1 = encabezados, fila 2+ = datos.
    Actualiza por número de serie, agrega nuevas.
    """
    resultados = {'hoja': nombre_hoja, 'importados': 0, 'actualizados': 0, 'errores': []}

    service = _get_service()
    filas   = _leer_hoja(service, sheet_id, nombre_hoja)

    if not filas or len(filas) < 2:
        resultados['errores'].append('La pestaña está vacía.')
        return resultados

    for i, fila in enumerate(filas[1:], start=2):
        try:
            if len(fila) < 3:
                continue

            aula    = str(fila[0]).strip() if len(fila) > 0 else ''
            curso   = str(fila[1]).strip() if len(fila) > 1 else ''
            n_serie = str(fila[2]).strip() if len(fila) > 2 else ''

            if not aula or not n_serie:
                continue

            pantalla = PantallaDigital.query.filter_by(numero_serie=n_serie).first()
            if pantalla:
                pantalla.aula          = aula
                pantalla.observaciones = f'Curso: {curso}' if curso else pantalla.observaciones
                resultados['actualizados'] += 1
            else:
                pantalla = PantallaDigital(
                    aula=aula, numero_serie=n_serie,
                    observaciones=f'Curso: {curso}' if curso else '',
                    estado='operativa'
                )
                db.session.add(pantalla)
                db.session.flush()
                h = HistorialPantalla(
                    pantalla_id=pantalla.id, evento='alta',
                    descripcion='Importada desde Google Sheets.', usuario='Sistema'
                )
                db.session.add(h)
                resultados['importados'] += 1

        except Exception as e:
            resultados['errores'].append(f'Fila {i}: {str(e)}')

    db.session.commit()
    return resultados


# ─────────────────────────────────────────────────────────────────────────────
#  IMPORTAR ALUMNOS
# ─────────────────────────────────────────────────────────────────────────────

def importar_alumnos(sheet_id, nombre_hoja):
    """
    Importa alumnos desde una pestaña de Google Sheets.
    Detecta el turno por el sufijo de la pestaña:
      - termina en ' M' → turno Mañana
      - termina en ' T' → turno Tarde
    El curso es el nombre sin el sufijo (ej: 'N1G1 M' → curso='N1G1', turno='M').

    Columnas: Apellido y Nombre | DNI
    Fila 1 = encabezados, fila 2+ = datos.
    Crea por DNI+turno si no existe, actualiza si ya existe.
    """
    resultados = {'hoja': nombre_hoja, 'importados': 0, 'actualizados': 0, 'errores': []}

    # Detectar turno y curso desde el nombre de la pestaña
    # Soporta: 'N1G1 TM', 'N1G4 TT', 'N1G1 M', 'N1G4 T'
    nombre_upper = nombre_hoja.strip().upper()
    if nombre_upper.endswith(' TM'):
        turno = 'M'
        curso = nombre_hoja.strip()[:-3].strip()
    elif nombre_upper.endswith(' TT'):
        turno = 'T'
        curso = nombre_hoja.strip()[:-3].strip()
    elif nombre_upper.endswith(' M'):
        turno = 'M'
        curso = nombre_hoja.strip()[:-2].strip()
    elif nombre_upper.endswith(' T'):
        turno = 'T'
        curso = nombre_hoja.strip()[:-2].strip()
    else:
        turno = 'M'
        curso = nombre_hoja.strip()

    service = _get_service()
    filas   = _leer_hoja(service, sheet_id, nombre_hoja)

    if not filas or len(filas) < 2:
        resultados['errores'].append('La pestaña está vacía.')
        return resultados

    for i, fila in enumerate(filas[1:], start=2):
        try:
            if len(fila) < 2:
                continue

            nombre_completo = fila[0].strip() if len(fila) > 0 else ''
            dni             = str(fila[1]).strip() if len(fila) > 1 else ''

            if not nombre_completo or not dni:
                continue

            if ',' in nombre_completo:
                partes   = nombre_completo.split(',', 1)
                apellido = partes[0].strip().title()
                nombre   = partes[1].strip().title()
            else:
                partes   = nombre_completo.split(' ', 1)
                apellido = partes[0].strip().title()
                nombre   = partes[1].strip().title() if len(partes) > 1 else ''

            alumno = Alumno.query.filter_by(dni=dni, turno=turno).first()
            if alumno:
                alumno.nombre   = nombre
                alumno.apellido = apellido
                alumno.curso    = curso
                alumno.turno    = turno
                resultados['actualizados'] += 1
            else:
                nuevo = Alumno(nombre=nombre, apellido=apellido, dni=dni,
                               curso=curso, turno=turno)
                db.session.add(nuevo)
                resultados['importados'] += 1

        except Exception as e:
            resultados['errores'].append(f'Fila {i}: {str(e)}')

    db.session.commit()
    return resultados


# ─────────────────────────────────────────────────────────────────────────────
#  IMPORTAR HORARIOS DE DOCENTES
# ─────────────────────────────────────────────────────────────────────────────

# Mapeo: código de módulo (texto) → número entero del dict MODULOS
CODIGO_A_NUMERO_MODULO = {
    'M1':    1,
    'M2':    2,
    'M3':    3,
    'M4':    4,
    'M5':    5,
    'M6':    6,
    'M7':    7,
    'M8':    8,
    'M8/T1': 8,
    'T1':    8,
    'T2':    9,
    'T3':    10,
    'T4':    11,
    'T5':    12,
    'T6':    13,
    'T7':    14,
    'T8':    15,
}

DIAS_NORMALIZADOS = {
    'LUNES':     'Lunes',
    'MARTES':    'Martes',
    'MIÉRCOLES': 'Miercoles',
    'MIERCOLES': 'Miercoles',
    'JUEVES':    'Jueves',
    'VIERNES':   'Viernes',
    'SABADO':    'Sabado',
    'SÁBADO':    'Sabado',
}


def importar_horarios_docentes(sheet_id, nombre_hoja='IMPORTAR'):
    """
    Lee la pestaña IMPORTAR de la planilla unificada y actualiza los horarios
    de cada docente mencionado en la base de datos.

    Estructura esperada (fila 1 = título, fila 2 = aviso,
    fila 3 = encabezados, fila 4+ = datos):

        apellido_nombre | dia       | modulo | materia  | curso
        RUBIO BRENDA    | MIÉRCOLES | T2     | HISTORIA | N1G6

    Por cada docente encontrado borra sus horarios actuales y los reemplaza.
    No crea docentes nuevos: si el nombre no existe en la BD lo registra como error.
    """
    from models_extra.horarios_notificaciones import HorarioDocente

    resultados = {
        'hoja':          nombre_hoja,
        'importados':    0,
        'actualizados':  0,
        'errores':       [],
        'advertencias':  [],
        'docentes_ok':   [],
        'docentes_no_encontrados': [],
    }

    service = _get_service()
    filas   = _leer_hoja(service, sheet_id, nombre_hoja)

    if not filas or len(filas) < 4:
        resultados['errores'].append(
            f'La pestaña "{nombre_hoja}" está vacía o no tiene datos suficientes.'
        )
        return resultados

    # Encabezados en fila 3 (índice 2)
    encabezados = [str(h).strip().lower() for h in filas[2]]
    for col in ['apellido_nombre', 'dia', 'modulo', 'materia', 'curso']:
        if col not in encabezados:
            resultados['errores'].append(
                f'Falta la columna "{col}". Encabezados encontrados: {encabezados}'
            )
            return resultados

    idx_nombre  = encabezados.index('apellido_nombre')
    idx_dia     = encabezados.index('dia')
    idx_modulo  = encabezados.index('modulo')
    idx_materia = encabezados.index('materia')
    idx_curso   = encabezados.index('curso')

    # Agrupar filas por docente
    filas_por_docente = {}
    for i, fila in enumerate(filas[3:], start=4):
        def get_col(idx):
            return str(fila[idx]).strip() if len(fila) > idx else ''

        nombre_docente = get_col(idx_nombre).upper()
        if not nombre_docente:
            continue

        if nombre_docente not in filas_por_docente:
            filas_por_docente[nombre_docente] = []

        filas_por_docente[nombre_docente].append({
            'fila':    i,
            'dia':     get_col(idx_dia).upper(),
            'modulo':  get_col(idx_modulo).upper(),
            'materia': get_col(idx_materia).upper(),
            'curso':   get_col(idx_curso).upper(),
        })

    if not filas_por_docente:
        resultados['errores'].append('No se encontraron filas con datos válidos.')
        return resultados

    # Procesar cada docente
    for nombre_docente, items in filas_por_docente.items():
        docente = _buscar_docente_por_nombre(nombre_docente)

        if not docente:
            resultados['docentes_no_encontrados'].append(nombre_docente)
            resultados['errores'].append(
                f'Docente no encontrado: "{nombre_docente}". '
                f'Verificá que el nombre coincida exactamente con el cargado en SIGA-Tec.'
            )
            continue

        # Borrar horarios actuales y reemplazar
        HorarioDocente.query.filter_by(docente_id=docente.id).delete()

        cargados  = 0
        ignorados = 0

        for item in items:
            dia_norm   = DIAS_NORMALIZADOS.get(item['dia'])
            num_modulo = CODIGO_A_NUMERO_MODULO.get(item['modulo'])

            if not dia_norm:
                resultados['advertencias'].append(
                    f'Fila {item["fila"]} ({nombre_docente}): '
                    f'día no reconocido "{item["dia"]}" — se ignora.'
                )
                ignorados += 1
                continue

            if not num_modulo:
                resultados['advertencias'].append(
                    f'Fila {item["fila"]} ({nombre_docente}): '
                    f'módulo no reconocido "{item["modulo"]}" — se ignora.'
                )
                ignorados += 1
                continue

            db.session.add(HorarioDocente(
                docente_id = docente.id,
                dia        = dia_norm,
                modulo     = num_modulo,
                materia    = item['materia'] or None,
                aula       = item['curso']   or None,
            ))
            cargados += 1

        # Actualizar materia y turno en el perfil del docente
        registros_cargados = [
            {'modulo': CODIGO_A_NUMERO_MODULO.get(i['modulo'], 0), 'materia': i['materia']}
            for i in items
            if DIAS_NORMALIZADOS.get(i['dia']) and CODIGO_A_NUMERO_MODULO.get(i['modulo'])
        ]
        _actualizar_materia_turno(docente, registros_cargados)

        resultados['importados']   += cargados
        resultados['actualizados'] += 1
        resultados['docentes_ok'].append(
            f'{nombre_docente} ({cargados} módulos'
            + (f', {ignorados} ignorados' if ignorados else '') + ')'
        )

    db.session.commit()
    return resultados


def _buscar_docente_por_nombre(nombre_completo):
    """
    Busca un docente activo en la BD a partir de 'APELLIDO NOMBRE'.
    Prueba separaciones de 1 y 2 palabras como apellido.
    Si hay un único docente con ese apellido, lo devuelve directamente.
    """
    partes = nombre_completo.strip().upper().split()
    if not partes:
        return None

    # Estrategia 1: primera palabra = apellido, resto = nombre
    if len(partes) >= 2:
        docente = Docente.query.filter(
            db.func.upper(Docente.apellido) == partes[0],
            db.func.upper(Docente.nombre)   == ' '.join(partes[1:]),
            Docente.activo == True
        ).first()
        if docente:
            return docente

    # Estrategia 2: primeras dos palabras = apellido, resto = nombre
    if len(partes) >= 3:
        docente = Docente.query.filter(
            db.func.upper(Docente.apellido) == ' '.join(partes[:2]),
            db.func.upper(Docente.nombre)   == ' '.join(partes[2:]),
            Docente.activo == True
        ).first()
        if docente:
            return docente

    # Estrategia 3: solo por apellido si hay resultado único
    docentes = Docente.query.filter(
        db.func.upper(Docente.apellido) == partes[0],
        Docente.activo == True
    ).all()
    if len(docentes) == 1:
        return docentes[0]

    return None


# ─────────────────────────────────────────────────────────────────────────────
#  PARSEAR EXCEL DE HORARIO INDIVIDUAL (subida directa desde SIGA-Tec)
# ─────────────────────────────────────────────────────────────────────────────

# Mapeo horario impreso → número de módulo
_HORARIO_A_MODULO = {
    '07.30 a 08.10': 1,
    '08.10 a 08.50': 2,
    '09.00 a 09.40': 3,
    '09.40 a 10.20': 4,
    '10.30 a 11.10': 5,
    '11.10 a 11.50': 6,
    '12.00 a 12.40': 7,
    '12.40 a 13.20': 8,
    '13.20 a 14.00': 9,
    '14.00 a 14.40': 10,
    '14.50 a 15.30': 11,
    '15.30 a 16.10': 12,
    '16.20 a 17.00': 13,
    '17.00 a 17.40': 14,
    '17.40 a 18.20': 15,
}

_DIAS_NORMALIZADOS = {
    'LUNES':     'Lunes',
    'MARTES':    'Martes',
    'MIÉRCOLES': 'Miercoles',
    'MIERCOLES': 'Miercoles',
    'JUEVES':    'Jueves',
    'VIERNES':   'Viernes',
    'SABADO':    'Sabado',
    'SÁBADO':    'Sabado',
}

# Mapeo código texto → número (para planilla unificada)
CODIGO_A_NUMERO_MODULO = {
    'M1': 1, 'M2': 2, 'M3': 3, 'M4': 4, 'M5': 5,
    'M6': 6, 'M7': 7, 'M8': 8, 'M8/T1': 8, 'T1': 8,
    'T2': 9, 'T3': 10, 'T4': 11, 'T5': 12,
    'T6': 13, 'T7': 14, 'T8': 15,
}

DIAS_NORMALIZADOS = _DIAS_NORMALIZADOS


def _normalizar_horario(horario):
    """
    Normaliza variantes de horario a formato '07.30 a 08.10' (con ceros).
    Soporta: '7.30 a 8.10', '07:30 a 08:10', '07:30 - 08:10', '7:30-8:10'.
    """
    import re
    if not horario:
        return ''
    h = str(horario).strip()
    # Reemplazar dos puntos por punto (08:10 → 08.10)
    h = re.sub(r'(\d{1,2}):(\d{2})', r'\1.\2', h)
    # Reemplazar separador de rango por ' a '
    h = re.sub(r'\s*[-–]\s*', ' a ', h)
    h = re.sub(r'\s+', ' ', h).strip()
    # Agregar cero inicial si falta: '7.30 a 8.10' → '07.30 a 08.10'
    def agregar_cero(match):
        return match.group(0).zfill(5)  # '7.30' → '07.30'
    h = re.sub(r'\b\d\.\d{2}\b', agregar_cero, h)
    return h


def _normalizar_curso(curso):
    """Convierte 'N 1 G 6' → 'N1G6'."""
    import re
    if not curso:
        return ''
    c = str(curso).strip().upper()
    c = re.sub(r'N\s*(\d+)\s*G\s*(\d+)', r'N\1G\2', c)
    return c


def _nombre_desde_archivo(nombre_archivo):
    """
    Extrae nombre del docente desde el nombre del archivo.
    'RUBIO_BRENDA_HORARIO_2026.xlsx' → 'RUBIO BRENDA'
    """
    import re
    stem = re.sub(r'\.xlsx?$', '', nombre_archivo, flags=re.IGNORECASE).upper()
    if '_HORARIO' in stem:
        nombre = stem[:stem.index('_HORARIO')].replace('_', ' ').strip()
        return nombre
    partes = stem.split('_')
    return ' '.join(partes[:2]).strip()


def _extraer_nombre_desde_planilla(todas_las_filas):
    """Intenta leer el nombre del docente desde la celda con label 'DOCENTE'."""
    for row in todas_las_filas[:6]:
        for i, cell in enumerate(row):
            if str(cell).strip().upper() == 'DOCENTE':
                for j in range(i + 1, len(row)):
                    val = str(row[j]).strip().upper() if row[j] else ''
                    if val and val not in ('CARGO', 'TURNO', 'TP1', 'TP2', 'TP3',
                                          'MAÑANA', 'TARDE', 'VESPERTINO', ''):
                        return val
    return None


def parsear_excel_horario(archivo_bytes, nombre_archivo, nombre_hoja='HORARIO 26'):
    """
    Parsea un archivo Excel de horario individual (formato ET N°7).
    Devuelve lista de dicts: [{dia, modulo, materia, curso}, ...]
    y el nombre del docente detectado.

    - archivo_bytes: contenido del archivo en bytes (de request.files)
    - nombre_archivo: nombre original del archivo (para extraer nombre docente)
    - nombre_hoja: nombre de la pestaña a leer
    """
    from io import BytesIO
    from openpyxl import load_workbook

    resultado = {
        'nombre_docente': None,
        'registros':      [],
        'advertencias':   [],
    }

    try:
        wb = load_workbook(BytesIO(archivo_bytes), read_only=True, data_only=True)
    except Exception as e:
        resultado['advertencias'].append(f'No se pudo abrir el archivo: {e}')
        return resultado

    # Elegir hoja
    if nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
    elif wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
    else:
        resultado['advertencias'].append('El archivo no tiene hojas.')
        return resultado

    todas_las_filas = list(ws.iter_rows(values_only=True))
    wb.close()

    # Nombre del docente — siempre desde el nombre del archivo (APELLIDO_NOMBRE_HORARIO...)
    # La planilla puede tener el nombre en formato 'NOMBRE APELLIDO' que no coincide con la BD
    nombre_docente = _nombre_desde_archivo(nombre_archivo)
    resultado['nombre_docente'] = nombre_docente

    # Encontrar fila de días
    fila_dias = None
    for idx, row in enumerate(todas_las_filas):
        vals = [str(c).strip().upper() if c else '' for c in row]
        if 'LUNES' in vals:
            fila_dias = idx
            break

    if fila_dias is None:
        resultado['advertencias'].append('No se encontró la grilla de días.')
        return resultado

    # Detectar columnas de inicio de cada día
    dia_cols = {}
    for col_idx, cell_val in enumerate(todas_las_filas[fila_dias]):
        dia = str(cell_val).strip().upper() if cell_val else ''
        if dia in _DIAS_NORMALIZADOS:
            dia_cols[dia] = col_idx

    if not dia_cols:
        resultado['advertencias'].append('No se detectaron columnas de días.')
        return resultado

    # Procesar filas de datos (desde fila_dias + 2 en adelante)
    # ultima_asig: lleva la última asignatura vista por día (para celdas mergeadas visualmente)
    ultima_asig = {dia: None for dia in dia_cols}

    for row in todas_las_filas[fila_dias + 2:]:
        if all(c is None for c in row):
            break

        for dia, col_inicio in dia_cols.items():
            try:
                asignatura_celda = str(row[col_inicio]).strip().upper() if len(row) > col_inicio and row[col_inicio] else ''
                horario          = _normalizar_horario(row[col_inicio + 1] if len(row) > col_inicio + 1 else '')
                curso            = _normalizar_curso(row[col_inicio + 2] if len(row) > col_inicio + 2 else '')
            except IndexError:
                continue

            # Carry: si la celda de asignatura está vacía, usar la última vista para este día
            if asignatura_celda:
                ultima_asig[dia] = asignatura_celda
            asignatura = ultima_asig[dia]

            if not asignatura:
                continue

            num_modulo = _HORARIO_A_MODULO.get(horario)
            if not num_modulo:
                if horario:
                    resultado['advertencias'].append(
                        f'Horario no reconocido: "{horario}" '
                        f'(día: {dia}, materia: {asignatura}) — se ignora.'
                    )
                continue

            resultado['registros'].append({
                'dia':     _DIAS_NORMALIZADOS[dia],
                'modulo':  num_modulo,
                'materia': asignatura,
                'curso':   curso,
            })

    return resultado



# Módulos de turno mañana (1-8) y tarde (9-15)
_MODULOS_MANANA = set(range(1, 9))
_MODULOS_TARDE  = set(range(9, 16))

def _actualizar_materia_turno(docente, registros):
    """
    Actualiza los campos materia y turno del docente
    a partir de los módulos importados.
    - materia: lista de materias únicas separadas por coma (excluye EXTRA CLASES)
    - turno: Mañana / Tarde / Mañana y Tarde
    """
    materias_unicas = []
    seen = set()
    for r in registros:
        m = (r.get('materia') or '').strip().upper()
        if m and m not in ('EXTRA CLASES', 'EXTRA', 'TAREAS DOCENTES', '') and m not in seen:
            seen.add(m)
            materias_unicas.append(m)

    modulos = {r['modulo'] for r in registros}
    tiene_manana = bool(modulos & _MODULOS_MANANA)
    tiene_tarde  = bool(modulos & _MODULOS_TARDE)

    if tiene_manana and tiene_tarde:
        turno = 'Mañana y Tarde'
    elif tiene_manana:
        turno = 'Mañana'
    elif tiene_tarde:
        turno = 'Tarde'
    else:
        turno = docente.turno  # no cambiar si no se puede determinar

    docente.materia = ', '.join(materias_unicas) if materias_unicas else docente.materia
    docente.turno   = turno

def importar_horarios_desde_excel(archivos):
    """
    Procesa una lista de archivos Excel subidos directamente desde SIGA-Tec
    y actualiza los horarios en la base de datos.

    - archivos: lista de dicts {'bytes': bytes, 'nombre': str}

    Devuelve el mismo formato de resultados que importar_horarios_docentes().
    """
    from models_extra.horarios_notificaciones import HorarioDocente

    resultados = {
        'importados':    0,
        'actualizados':  0,
        'errores':       [],
        'advertencias':  [],
        'docentes_ok':   [],
        'docentes_no_encontrados': [],
    }

    for archivo in archivos:
        parsed = parsear_excel_horario(archivo['bytes'], archivo['nombre'])

        # Agregar advertencias del parseo
        for adv in parsed['advertencias']:
            resultados['advertencias'].append(f"{archivo['nombre']}: {adv}")

        nombre_docente = parsed['nombre_docente']
        registros      = parsed['registros']

        if not registros:
            resultados['errores'].append(
                f"{archivo['nombre']}: no se encontraron módulos válidos."
            )
            continue

        # Buscar docente en la BD
        docente = _buscar_docente_por_nombre(nombre_docente)
        if not docente:
            resultados['docentes_no_encontrados'].append(nombre_docente)
            resultados['errores'].append(
                f'Docente no encontrado: "{nombre_docente}". '
                f'Verificá que el nombre del archivo coincida con el cargado en SIGA-Tec '
                f'(formato: APELLIDO_NOMBRE_HORARIO_2026.xlsx).'
            )
            continue

        # ACUMULAR: no borramos los módulos existentes.
        # Si el docente tiene turno Varios con 2 planillas, la segunda se suma a la primera.
        # Solo actualizamos si el slot exacto (dia + modulo + curso) ya existe.
        for item in registros:
            existente = HorarioDocente.query.filter_by(
                docente_id=docente.id,
                dia=item['dia'],
                modulo=item['modulo'],
                aula=item['curso'] or None,
            ).first()
            if existente:
                existente.materia = item['materia'] or existente.materia
            else:
                db.session.add(HorarioDocente(
                    docente_id=docente.id,
                    dia=item['dia'],
                    modulo=item['modulo'],
                    materia=item['materia'] or None,
                    aula=item['curso'] or None,
                ))

        # Actualizar materia y turno en el perfil del docente
        _actualizar_materia_turno(docente, registros)

        resultados['importados']   += len(registros)
        resultados['actualizados'] += 1
        resultados['docentes_ok'].append(
            f'{nombre_docente} ({len(registros)} módulos)'
        )

    db.session.commit()
    return resultados


def importar_horarios_docentes(sheet_id, nombre_hoja='IMPORTAR'):
    """
    Lee la pestaña IMPORTAR de la planilla unificada y actualiza los horarios
    de cada docente mencionado en la base de datos.

    Estructura esperada (fila 1 = título, fila 2 = aviso,
    fila 3 = encabezados, fila 4+ = datos):

        apellido_nombre | dia       | modulo | materia  | curso
        RUBIO BRENDA    | MIÉRCOLES | T2     | HISTORIA | N1G6

    Por cada docente encontrado borra sus horarios actuales y los reemplaza.
    No crea docentes nuevos: si el nombre no existe en la BD lo registra como error.
    """
    from models_extra.horarios_notificaciones import HorarioDocente

    resultados = {
        'hoja':          nombre_hoja,
        'importados':    0,
        'actualizados':  0,
        'errores':       [],
        'advertencias':  [],
        'docentes_ok':   [],
        'docentes_no_encontrados': [],
    }

    service = _get_service()
    filas   = _leer_hoja(service, sheet_id, nombre_hoja)

    if not filas or len(filas) < 4:
        resultados['errores'].append(
            f'La pestaña "{nombre_hoja}" está vacía o no tiene datos suficientes.'
        )
        return resultados

    # Encabezados en fila 3 (índice 2)
    encabezados = [str(h).strip().lower() for h in filas[2]]
    for col in ['apellido_nombre', 'dia', 'modulo', 'materia', 'curso']:
        if col not in encabezados:
            resultados['errores'].append(
                f'Falta la columna "{col}". Encabezados encontrados: {encabezados}'
            )
            return resultados

    idx_nombre  = encabezados.index('apellido_nombre')
    idx_dia     = encabezados.index('dia')
    idx_modulo  = encabezados.index('modulo')
    idx_materia = encabezados.index('materia')
    idx_curso   = encabezados.index('curso')

    # Agrupar filas por docente
    filas_por_docente = {}
    for i, fila in enumerate(filas[3:], start=4):
        def get_col(idx):
            return str(fila[idx]).strip() if len(fila) > idx else ''

        nombre_docente = get_col(idx_nombre).upper()
        if not nombre_docente:
            continue

        if nombre_docente not in filas_por_docente:
            filas_por_docente[nombre_docente] = []

        filas_por_docente[nombre_docente].append({
            'fila':    i,
            'dia':     get_col(idx_dia).upper(),
            'modulo':  get_col(idx_modulo).upper(),
            'materia': get_col(idx_materia).upper(),
            'curso':   get_col(idx_curso).upper(),
        })

    if not filas_por_docente:
        resultados['errores'].append('No se encontraron filas con datos válidos.')
        return resultados

    for nombre_docente, items in filas_por_docente.items():
        docente = _buscar_docente_por_nombre(nombre_docente)

        if not docente:
            resultados['docentes_no_encontrados'].append(nombre_docente)
            resultados['errores'].append(
                f'Docente no encontrado: "{nombre_docente}". '
                f'Verificá que el nombre coincida exactamente con el cargado en SIGA-Tec.'
            )
            continue

        HorarioDocente.query.filter_by(docente_id=docente.id).delete()

        cargados  = 0
        ignorados = 0

        for item in items:
            dia_norm   = DIAS_NORMALIZADOS.get(item['dia'])
            num_modulo = CODIGO_A_NUMERO_MODULO.get(item['modulo'])

            if not dia_norm:
                resultados['advertencias'].append(
                    f'Fila {item["fila"]} ({nombre_docente}): '
                    f'día no reconocido "{item["dia"]}" — se ignora.'
                )
                ignorados += 1
                continue

            if not num_modulo:
                resultados['advertencias'].append(
                    f'Fila {item["fila"]} ({nombre_docente}): '
                    f'módulo no reconocido "{item["modulo"]}" — se ignora.'
                )
                ignorados += 1
                continue

            db.session.add(HorarioDocente(
                docente_id = docente.id,
                dia        = dia_norm,
                modulo     = num_modulo,
                materia    = item['materia'] or None,
                aula       = item['curso']   or None,
            ))
            cargados += 1

        # Actualizar materia y turno en el perfil del docente
        registros_cargados = [
            {'modulo': CODIGO_A_NUMERO_MODULO.get(i['modulo'], 0), 'materia': i['materia']}
            for i in items
            if DIAS_NORMALIZADOS.get(i['dia']) and CODIGO_A_NUMERO_MODULO.get(i['modulo'])
        ]
        _actualizar_materia_turno(docente, registros_cargados)

        resultados['importados']   += cargados
        resultados['actualizados'] += 1
        resultados['docentes_ok'].append(
            f'{nombre_docente} ({cargados} módulos'
            + (f', {ignorados} ignorados' if ignorados else '') + ')'
        )

    db.session.commit()
    return resultados


def _buscar_docente_por_nombre(nombre_completo):
    """
    Busca un docente activo por 'APELLIDO NOMBRE'.
    Prueba separaciones de 1 y 2 palabras como apellido.
    """
    partes = nombre_completo.strip().upper().split()
    if not partes:
        return None

    if len(partes) >= 2:
        docente = Docente.query.filter(
            db.func.upper(Docente.apellido) == partes[0],
            db.func.upper(Docente.nombre)   == ' '.join(partes[1:]),
            Docente.activo == True
        ).first()
        if docente:
            return docente

    if len(partes) >= 3:
        docente = Docente.query.filter(
            db.func.upper(Docente.apellido) == ' '.join(partes[:2]),
            db.func.upper(Docente.nombre)   == ' '.join(partes[2:]),
            Docente.activo == True
        ).first()
        if docente:
            return docente

    docentes = Docente.query.filter(
        db.func.upper(Docente.apellido) == partes[0],
        Docente.activo == True
    ).all()
    if len(docentes) == 1:
        return docentes[0]

    return None
