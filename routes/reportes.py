from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from models import db, PrestamoCarro, PrestamoNetbook, Docente, Carro
from datetime import datetime, timedelta

reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')


def _filtrar_prestamos_carros(periodo, fecha_desde=None, fecha_hasta=None):
    ahora = datetime.utcnow()
    query = PrestamoCarro.query

    if fecha_desde and fecha_hasta:
        try:
            d_desde  = datetime.strptime(fecha_desde, '%Y-%m-%d')
            d_hasta  = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query    = query.filter(PrestamoCarro.hora_retiro >= d_desde,
                                    PrestamoCarro.hora_retiro <= d_hasta)
        except ValueError:
            pass
    elif fecha_desde:
        try:
            d_desde = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query   = query.filter(PrestamoCarro.hora_retiro >= d_desde)
        except ValueError:
            pass
    elif periodo == 'hoy':
        query = query.filter(PrestamoCarro.hora_retiro >= ahora.replace(hour=0, minute=0, second=0))
    elif periodo == 'semana':
        query = query.filter(PrestamoCarro.hora_retiro >= ahora - timedelta(days=7))
    elif periodo == 'mes':
        query = query.filter(PrestamoCarro.hora_retiro >= ahora - timedelta(days=30))

    return query.order_by(PrestamoCarro.hora_retiro.desc()).all()


def _filtrar_prestamos_netbooks(periodo, fecha_desde=None, fecha_hasta=None):
    ahora = datetime.utcnow()
    query = PrestamoNetbook.query

    if fecha_desde and fecha_hasta:
        try:
            d_desde = datetime.strptime(fecha_desde, '%Y-%m-%d')
            d_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query   = query.filter(PrestamoNetbook.hora_retiro >= d_desde,
                                   PrestamoNetbook.hora_retiro <= d_hasta)
        except ValueError:
            pass
    elif fecha_desde:
        try:
            d_desde = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query   = query.filter(PrestamoNetbook.hora_retiro >= d_desde)
        except ValueError:
            pass
    elif periodo == 'hoy':
        query = query.filter(PrestamoNetbook.hora_retiro >= ahora.replace(hour=0, minute=0, second=0))
    elif periodo == 'semana':
        query = query.filter(PrestamoNetbook.hora_retiro >= ahora - timedelta(days=7))
    elif periodo == 'mes':
        query = query.filter(PrestamoNetbook.hora_retiro >= ahora - timedelta(days=30))

    return query.order_by(PrestamoNetbook.hora_retiro.desc()).all()


@reportes_bp.route('/')
@login_required
def index():
    if not current_user.tiene_permiso('reportes'):
        flash('Credenciales no válidas para acceder a reportes.', 'danger')
        return redirect(url_for('main.dashboard'))
    return render_template('reportes/index.html')


@reportes_bp.route('/estadisticas')
@login_required
def estadisticas():
    if not current_user.tiene_permiso('estadisticas'):
        flash('Credenciales no válidas para ver estadísticas.', 'danger')
        return redirect(url_for('main.dashboard'))

    from sqlalchemy import func
    top_docentes = db.session.query(
        Docente, func.count(PrestamoCarro.id).label('total')
    ).join(PrestamoCarro).group_by(Docente.id).order_by(
        func.count(PrestamoCarro.id).desc()).limit(10).all()

    top_materias = db.session.query(
        Docente.materia, func.count(PrestamoCarro.id).label('total')
    ).join(PrestamoCarro).group_by(Docente.materia).order_by(
        func.count(PrestamoCarro.id).desc()).limit(10).all()

    return render_template('reportes/estadisticas.html',
                           top_docentes=top_docentes,
                           top_materias=top_materias)


# ─────────────────────────────────────────────────────────────────────────────
#  PDFs
# ─────────────────────────────────────────────────────────────────────────────

@reportes_bp.route('/pdf/carros')
@login_required
def pdf_carros():
    from services.pdf_reportes import pdf_listado_carros
    carros = Carro.query.filter(Carro.estado != 'baja').order_by(Carro.numero_fisico).all()
    return pdf_listado_carros(carros)


@reportes_bp.route('/pdf/carro/<int:id>')
@login_required
def pdf_netbooks_carro(id):
    from services.pdf_reportes import pdf_netbooks_por_carro
    carro = Carro.query.get_or_404(id)
    return pdf_netbooks_por_carro(carro)


@reportes_bp.route('/pdf/asignadas')
@login_required
def pdf_asignadas():
    from services.pdf_reportes import pdf_netbooks_asignadas
    carro_id = request.args.get('carro_id', type=int)
    if carro_id:
        carro = Carro.query.get_or_404(carro_id)
        return pdf_netbooks_asignadas(carro)
    return pdf_netbooks_asignadas()


@reportes_bp.route('/pdf/servicio-tecnico')
@login_required
def pdf_servicio():
    from services.pdf_reportes import pdf_servicio_tecnico
    return pdf_servicio_tecnico()


@reportes_bp.route('/pdf/historial-carros')
@login_required
def pdf_hist_carros():
    from services.pdf_reportes import pdf_historial_carros
    periodo     = request.args.get('periodo', 'todos')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    prestamos   = _filtrar_prestamos_carros(periodo, fecha_desde, fecha_hasta)
    label       = f'{fecha_desde} al {fecha_hasta}' if fecha_desde else periodo
    return pdf_historial_carros(prestamos, label)


@reportes_bp.route('/pdf/historial-netbooks')
@login_required
def pdf_hist_netbooks():
    from services.pdf_reportes import pdf_historial_netbooks
    periodo     = request.args.get('periodo', 'todos')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    prestamos   = _filtrar_prestamos_netbooks(periodo, fecha_desde, fecha_hasta)
    label       = f'{fecha_desde} al {fecha_hasta}' if fecha_desde else periodo
    return pdf_historial_netbooks(prestamos, label)


@reportes_bp.route('/pdf/estadisticas')
@login_required
def pdf_estadisticas():
    if not current_user.tiene_permiso('estadisticas'):
        flash('Credenciales no válidas.', 'danger')
        return redirect(url_for('main.dashboard'))
    from services.pdf_reportes import pdf_estadisticas
    from sqlalchemy import func
    top_docentes = db.session.query(
        Docente, func.count(PrestamoCarro.id).label('total')
    ).join(PrestamoCarro).group_by(Docente.id).order_by(
        func.count(PrestamoCarro.id).desc()).limit(10).all()
    top_materias = db.session.query(
        Docente.materia, func.count(PrestamoCarro.id).label('total')
    ).join(PrestamoCarro).group_by(Docente.materia).order_by(
        func.count(PrestamoCarro.id).desc()).limit(10).all()
    return pdf_estadisticas(top_docentes, top_materias)


@reportes_bp.route('/pdf/asignaciones/carro/<int:carro_id>')
@login_required
def pdf_asignaciones_carro(carro_id):
    """PDF con todas las netbooks del carro y sus alumnos asignados."""
    from services.pdf_reportes import generar_pdf_asignaciones_carro
    from flask import send_file
    carro  = Carro.query.get_or_404(carro_id)
    buffer = generar_pdf_asignaciones_carro(carro)
    nombre = f'asignaciones_carro_{carro.numero_fisico}.pdf'
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=nombre)


@reportes_bp.route('/pdf/asignaciones/netbook/<int:netbook_id>')
@login_required
def pdf_asignacion_netbook(netbook_id):
    """PDF de una netbook individual con su alumno asignado."""
    from services.pdf_reportes import generar_pdf_asignacion_netbook
    from models import Netbook
    from flask import send_file
    netbook = Netbook.query.get_or_404(netbook_id)
    buffer  = generar_pdf_asignacion_netbook(netbook)
    nombre  = f'asignacion_netbook_{netbook.numero_interno}.pdf'
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=nombre)

@reportes_bp.route('/pdf/inventario/carro/<int:carro_id>')
@login_required
def pdf_inventario_carro(carro_id):
    """PDF simple con N° interno y N° serie de todas las netbooks del carro."""
    from services.pdf_reportes import pdf_inventario_carro
    from flask import send_file
    carro  = Carro.query.get_or_404(carro_id)
    buffer = pdf_inventario_carro(carro)
    nombre = f'inventario_carro_{carro.numero_fisico}.pdf'
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=nombre)


# ─────────────────────────────────────────────────────────────────────────────
#  INVENTARIO INTEGRAL DE NETBOOKS
# ─────────────────────────────────────────────────────────────────────────────

@reportes_bp.route('/inventario-netbooks/pdf')
@login_required
def inventario_netbooks_pdf():
    """PDF landscape: todas las netbooks del sistema + asignaciones internas."""
    if not current_user.tiene_permiso('reportes'):
        flash('Credenciales no válidas para acceder a reportes.', 'danger')
        return redirect(url_for('main.dashboard'))
    from services.pdf_reportes import pdf_inventario_integral_netbooks
    return pdf_inventario_integral_netbooks()


@reportes_bp.route('/inventario-netbooks/excel')
@login_required
def inventario_netbooks_excel():
    """Excel con 2 hojas: Netbooks en Carros + Asignaciones Internas."""
    if not current_user.tiene_permiso('reportes'):
        flash('Credenciales no válidas para acceder a reportes.', 'danger')
        return redirect(url_for('main.dashboard'))

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file
    from datetime import timedelta
    from models import Netbook, AsignacionInterna

    ARG_OFFSET = timedelta(hours=-3)
    ahora_arg  = (datetime.utcnow() + ARG_OFFSET).strftime('%d/%m/%Y %H:%M')

    wb = openpyxl.Workbook()

    AZUL  = '1A2A6C'
    VERDE = '16A34A'
    NARAN = 'E8821A'
    ROJO  = 'DC2626'
    GRIS  = 'F3F4F6'
    BLANC = 'FFFFFF'

    def _fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def _font(bold=False, color='000000', size=10):
        return Font(bold=bold, color=color, size=size, name='Calibri')

    def _border():
        thin = Side(style='thin', color='D1D5DB')
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _left():
        return Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ── HOJA 1: Netbooks en Carros ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Netbooks en Carros'
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells('A1:H1')
    c = ws1['A1']
    c.value     = 'INVENTARIO INTEGRAL DE NETBOOKS — E.T. N°7 D.E. 5'
    c.font      = _font(bold=True, color=BLANC, size=13)
    c.fill      = _fill(AZUL)
    c.alignment = _center()
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells('A2:H2')
    c = ws1['A2']
    c.value     = f'Netbooks en Carros  ·  Generado: {ahora_arg}'
    c.font      = _font(color='6B7280', size=9)
    c.alignment = _center()
    ws1.row_dimensions[2].height = 16
    ws1.row_dimensions[3].height = 6

    HEADERS_NB = ['Carro', 'División', 'Aula', 'N° Interno',
                  'N° Serie', 'Alumno Mañana', 'Alumno Tarde', 'Estado']
    for col, h in enumerate(HEADERS_NB, start=1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font      = _font(bold=True, color=BLANC, size=10)
        c.fill      = _fill(AZUL)
        c.alignment = _center()
        c.border    = _border()
    ws1.row_dimensions[4].height = 20

    netbooks = (Netbook.query
                .join(Carro)
                .order_by(Carro.numero_fisico, Netbook.numero_interno)
                .all())

    def _nom_alumno(nb, turno):
        obj = nb.alumno_manana if turno == 'M' else nb.alumno_tarde
        if obj:
            return f'{obj.apellido}, {obj.nombre}'
        if turno == 'M' and nb.alumno:
            return nb.alumno
        return '—'

    total_op = total_st = total_bja = 0

    for fila, nb in enumerate(netbooks, start=5):
        bg = _fill(GRIS) if fila % 2 == 0 else _fill(BLANC)
        carro = nb.carro

        if nb.estado == 'operativa':
            estado_txt = 'Operativa';        total_op  += 1
        elif nb.estado == 'servicio_tecnico':
            estado_txt = 'Servicio Técnico'; total_st  += 1
        else:
            estado_txt = nb.estado.capitalize(); total_bja += 1

        valores = [
            carro.display if carro else '—',
            carro.division or '—' if carro else '—',
            carro.aula     or '—' if carro else '—',
            nb.numero_interno or '—',
            nb.numero_serie   or '—',
            _nom_alumno(nb, 'M'),
            _nom_alumno(nb, 'T'),
            estado_txt,
        ]
        for col, val in enumerate(valores, start=1):
            c = ws1.cell(row=fila, column=col, value=val)
            c.font      = _font(size=9)
            c.fill      = bg
            c.border    = _border()
            c.alignment = _center() if col in (1, 3, 4, 8) else _left()

        c_est = ws1.cell(row=fila, column=8)
        if nb.estado == 'servicio_tecnico':
            c_est.font = _font(bold=True, color=NARAN, size=9)
        elif nb.estado not in ('operativa', 'servicio_tecnico'):
            c_est.font = _font(bold=True, color=ROJO,  size=9)
        else:
            c_est.font = _font(bold=True, color=VERDE, size=9)

    fila_res = len(netbooks) + 5
    ws1.merge_cells(f'A{fila_res}:G{fila_res}')
    c = ws1.cell(row=fila_res, column=1,
                 value=f'TOTAL: {len(netbooks)}  |  Operativas: {total_op}  |  '
                       f'Servicio Técnico: {total_st}  |  Baja: {total_bja}')
    c.font      = _font(bold=True, color=BLANC, size=9)
    c.fill      = _fill(AZUL)
    c.alignment = _left()
    c.border    = _border()
    ws1.cell(row=fila_res, column=8).fill   = _fill(AZUL)
    ws1.cell(row=fila_res, column=8).border = _border()
    ws1.row_dimensions[fila_res].height = 18

    for i, ancho in enumerate([10, 16, 12, 11, 28, 28, 28, 14], start=1):
        ws1.column_dimensions[get_column_letter(i)].width = ancho
    ws1.freeze_panes = 'A5'

    # ── HOJA 2: Asignaciones Internas ─────────────────────────────────────────
    ws2 = wb.create_sheet('Asignaciones Internas')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:F1')
    c = ws2['A1']
    c.value     = 'ASIGNACIONES INTERNAS ACTIVAS — E.T. N°7 D.E. 5'
    c.font      = _font(bold=True, color=BLANC, size=13)
    c.fill      = _fill(VERDE)
    c.alignment = _center()
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells('A2:F2')
    c = ws2['A2']
    c.value     = f'Netbooks asignadas a docentes o áreas  ·  {ahora_arg}'
    c.font      = _font(color='6B7280', size=9)
    c.alignment = _center()
    ws2.row_dimensions[2].height = 16
    ws2.row_dimensions[3].height = 6

    HEADERS_AI = ['N° Interno', 'N° Serie', 'Modelo',
                  'Destinatario / Área', 'Motivo', 'Fecha Asignación']
    for col, h in enumerate(HEADERS_AI, start=1):
        c = ws2.cell(row=4, column=col, value=h)
        c.font      = _font(bold=True, color=BLANC, size=10)
        c.fill      = _fill(VERDE)
        c.alignment = _center()
        c.border    = _border()
    ws2.row_dimensions[4].height = 20

    asignaciones = (AsignacionInterna.query
                    .filter_by(activa=True)
                    .order_by(AsignacionInterna.id)
                    .all())

    for fila, a in enumerate(asignaciones, start=5):
        bg = _fill(GRIS) if fila % 2 == 0 else _fill(BLANC)
        fecha_asig = ((a.fecha_asignacion + ARG_OFFSET).strftime('%d/%m/%Y')
                      if a.fecha_asignacion else '—')
        valores = [
            a.numero_interno or '—',
            a.numero_serie   or '—',
            a.modelo         or '—',
            a.destinatario,
            a.motivo         or '—',
            fecha_asig,
        ]
        for col, val in enumerate(valores, start=1):
            c = ws2.cell(row=fila, column=col, value=val)
            c.font      = _font(size=9)
            c.fill      = bg
            c.border    = _border()
            c.alignment = _center() if col in (1, 6) else _left()

    fila_res2 = len(asignaciones) + 5
    ws2.merge_cells(f'A{fila_res2}:F{fila_res2}')
    c = ws2.cell(row=fila_res2, column=1,
                 value=f'TOTAL asignaciones internas activas: {len(asignaciones)}')
    c.font      = _font(bold=True, color=BLANC, size=9)
    c.fill      = _fill(VERDE)
    c.alignment = _left()
    c.border    = _border()
    ws2.row_dimensions[fila_res2].height = 18

    for i, ancho in enumerate([12, 26, 30, 32, 28, 16], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = ancho
    ws2.freeze_panes = 'A5'

    # ── Enviar ────────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f'inventario_netbooks_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

